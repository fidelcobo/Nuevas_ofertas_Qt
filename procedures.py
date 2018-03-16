
from PyQt5 import QtGui, QtWidgets, QtCore
from datetime import datetime
from listas import list_uptime_codes, list_uptime_descr
import openpyxl
import os
import csv


import locale

locale.setlocale(locale.LC_ALL, 'FR')


def busca_columnas(sheet: object, lista_busca: list, fila_busca: str) -> object:

    columnas = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T',
                'U','V','W','X','Y','Z', 'AA','AB','AC','AD','AE']

    result_busca = []

    for items in lista_busca:
        for col in columnas:
            cell = col + fila_busca
            if sheet[cell].value == items:
                result_busca.append(col)
                break
    if len(result_busca) == len(lista_busca):  # Se han encontrado todos los campos relevantes
        ok = True
    else:
        ok = False

    return ok, result_busca


def get_ultima_fila(hoja, col_code):
    fila = 11
    fin = False
    while not fin:
        cell = col_code + str(fila)
        if not hoja[cell].value:
            fin = True
        else:
            fila += 1
            fin = False
    return (fila-1)


def csv_from_excel(entrada, salida, instance):

    # VARIANTE CON xlrd. Pone los números de línea en float
    # print(entrada)
    # with xlrd.open_workbook(entrada) as wb:
    #     sh = wb.sheet_by_index(0)  # or wb.sheet_by_name('name_of_the_sheet_here')
    #     with open(salida, 'w', newline='') as f:
    #         c = csv.writer(f, delimiter=';', quoting=csv.QUOTE_MINIMAL)
    #         for r in range(sh.nrows):
    #             c.writerow(sh.row_values(r))
    ok = False

    while not ok:
        try:
            wb = openpyxl.load_workbook(entrada)
            sh = wb.get_active_sheet()  # or wb.sheet_by_name('name_of_the_sheet_here')

            with open(salida, 'w', newline='') as f:
                c = csv.writer(f, dialect='excel', delimiter=';')
                for r in sh.rows:
                    c.writerow([cell.value for cell in r])
            ok = True

        except PermissionError:
            QtWidgets.QMessageBox.warning(instance, 'Procesado de oferta',
                                          'Fichero csv de destino {} ya abierto \n'
                                          'por favor ciérrelo y pulse Aceptar'.format(salida))


def pass_to_excel(lista: [], instance):

    actual_dir = os.path.dirname(__file__)

    # Ahora abrimos el fichero Excel auxiliar de plantilla de MS
    filen = os.path.join(actual_dir, 'plantilla_ms.xlsx')  # Este fichero no se toca. Hace de plantilla
    sheet_name = 'ms'
    if not os.path.exists(filen):
        instance.QtWidgets.QMessageBox.critical(instance, 'Procesado de oferta',
                                                'El fichero {} no se encuentra\n'.format(filen))
        return None

    try:
        libro = openpyxl.load_workbook(filen)
        hoja = libro.get_sheet_by_name(sheet_name)

    except PermissionError:
        instance.QtWidgets.QMessageBox.critical(instance, 'Procesado de oferta',
                                               'Fichero Excel {} no se puede abrir\n'
                                               'Compruebe que no está ya abierto'.format(filen))

        return None

    curr_row = 2
    num_fila = 10

    for reg in lista:
        fila = str(curr_row)
        fila_sig = str(curr_row + 1)

        hoja['A' + fila] = num_fila
        hoja['A' + fila_sig] = num_fila + 1
        hoja['B' + fila_sig] = num_fila
        hoja['H' + fila] = 2
        hoja['H' + fila_sig] = 20
        hoja['C' + fila] = 'Dimension Data'
        hoja['C' + fila_sig] = reg.manufacturer
        hoja['E' + fila] = reg.unit
        hoja['E' + fila_sig] = reg.unit
        hoja['F' + fila] = reg.sku_uptime
        hoja['F' + fila_sig] = reg.backout_name
        hoja['G' + fila] = reg.descr_uptime
        hoja['G' + fila_sig] = reg.code
        hoja['I' + fila] = reg.code
        hoja['I' + fila_sig] = 0
        hoja['J' + fila] = reg.manufacturer
        hoja['J' + fila_sig] = ''
        hoja['K' + fila] = 1
        hoja['K' + fila_sig] = 1
        hoja['L' + fila] = 'EA'
        hoja['L' + fila_sig] = 'EA'
        hoja['M' + fila] = 'EUR'
        hoja['M' + fila_sig] = 'EUR'
        hoja['N' + fila] = 'Fixed'
        hoja['N' + fila_sig] = 'Fixed'
        hoja['O' + fila] = reg.list_price_back
        hoja['O' + fila_sig] = reg.list_price_back
        hoja['P' + fila] = locale.format('%10.2f', float(reg.venta_mant))
        hoja['P' + fila_sig] = reg.list_price_back
        hoja['Q' + fila] = locale.format('%10.2f', float(reg.cost_unit_manten))
        hoja['Q' + fila_sig] = locale.format('%10.2f', float(reg.coste_unit_back))
        init_date_formateado = datetime.strptime(reg.init_date,"%d/%m/%Y")
        end_date_formateado = datetime.strptime(reg.end_date,"%d/%m/%Y")
        fecha_init = '{}/{}/{}'.format(init_date_formateado.day, init_date_formateado.month, init_date_formateado.year)
        fecha_init_limpia = '{}{}{}'.format(str(init_date_formateado.year), str(init_date_formateado.month).zfill(2),
                                            str(init_date_formateado.day).zfill(2))
        fecha_fin = '{}/{}/{}'.format(end_date_formateado.day, end_date_formateado.month, end_date_formateado.year)
        hoja['X' + fila] = fecha_init
        hoja['X' + fila_sig] = fecha_init
        hoja['Y' + fila] = fecha_fin
        hoja['Y' + fila_sig] = fecha_fin
        hoja['AA' + fila] = ('StartDate=' + fecha_init_limpia + '#Duration=' + str(reg.durac) +
                            '#InvoiceInterval=Yearly#InvoiceMode=anticipated')
        hoja['AA' + fila_sig] = ('StartDate=' + fecha_init_limpia + '#Duration=' + str(reg.durac) +
                            '#InvoiceInterval=Yearly#InvoiceMode=anticipated')

        hoja['AF' + fila] = reg.tech
        hoja['AF' + fila_sig] = reg.tech
        curr_row += 2
        num_fila += 3

    return libro


def clasificar_articulos(lista_articulos, instance):

    lista_checkpoint = [p
                        for p in lista_articulos
                        if p.manufacturer == 'Checkpoint']
    instance.lista_checkpoint = lista_checkpoint

    lista_fortinet = [p
                      for p in lista_articulos
                      if p.manufacturer == 'Fortinet']
    instance.lista_fortinet = lista_fortinet

    lista_hp = [p
                for p in lista_articulos
                if p.manufacturer == 'HP']
    instance.lista_hp = lista_hp

    lista_f5 = [p
                for p in lista_articulos
                if p.manufacturer == 'F5 Networks']
    instance.lista_f5 = lista_f5

    lista_cisco = [p
                   for p in lista_articulos
                   if p.manufacturer == 'CISCO']
    instance.lista_cisco = lista_cisco

    lista_alcatel = [p
                     for p in lista_articulos
                     if p.manufacturer == 'Alcatel']
    instance.lista_alcatel = lista_alcatel

    lista_riverbed = [p
                      for p in lista_articulos
                      if p.manufacturer == 'Riverbed']
    instance.lista_riverbed = lista_riverbed

    lista_paloalto = [p
                      for p in lista_articulos
                      if p.manufacturer == 'Palo Alto Networks']
    instance.lista_paloalto = lista_paloalto

    lista_juniper = [p
                     for p in lista_articulos
                     if p.manufacturer == 'Juniper']
    instance.lista_juniper = lista_juniper

    lista_bluecoat = [p
                      for p in lista_articulos
                      if p.manufacturer == 'Bluecoat']
    instance.lista_bluecoat = lista_bluecoat

    lista_brocade = [p
                     for p in lista_articulos
                     if p.manufacturer == 'Brocade']
    instance.lista_brocade = lista_brocade


def diff_days(init_date, end_date):  # Calcula la diferencia en días entre dos fechas dadas en formato string AA/MM/DD

    init_date = datetime.strptime(init_date,"%d/%m/%Y")
    end_date = datetime.strptime(end_date,"%d/%m/%Y")
    diff = end_date - init_date
    return diff.days


def hacer_oferta(directorio, instance):

    oferta_productos('Checkpoint', instance.lista_checkpoint, instance, directorio)
    oferta_productos('Fortinet', instance.lista_fortinet, instance, directorio)
    oferta_productos('F5', instance.lista_f5, instance, directorio)
    oferta_productos('HP', instance.lista_hp, instance, directorio)
    oferta_productos('Riverbed', instance.lista_riverbed, instance, directorio)
    oferta_productos('Palo Alto Networks', instance.lista_paloalto, instance, directorio)
    oferta_productos('Brocade', instance.lista_brocade, instance, directorio)
    oferta_productos('Juniper', instance.lista_juniper, instance, directorio)
    oferta_productos('Bluecoat', instance.lista_bluecoat, instance, directorio)
    oferta_productos('Alcatel', instance.lista_alcatel, instance, directorio)
    oferta_productos('CISCO', instance.lista_cisco, instance, directorio)


def oferta_productos(fabr, lista, instance, directorio):

    if lista:
        curr_row = 2
        num_fila = 10
        actual_dir = os.path.dirname(__file__)
        filen = os.path.join(actual_dir, 'plantilla_prod.xlsx')  # Este fichero no se toca. Hace de plantilla
        sheet_name = 'BOM'
        libro = openpyxl.load_workbook(filen)
        hoja = libro.get_sheet_by_name(sheet_name)

        for items in lista:
            fila = str(curr_row)
            hoja['A' + fila] = num_fila
            hoja['C' + fila] = items.manufacturer
            hoja['H' + fila] = 1
            hoja['L' + fila] = 'EA'
            hoja['M' + fila] = 'EUR'
            hoja['N' + fila] = 'Fixed'
            hoja['F' + fila] = items.code
            hoja['G' + fila] = items.descripcion_prod
            hoja['K' + fila] = items.unit
            hoja['O' + fila] = locale.format('%10.2f', items.list_price)
            hoja['Q' + fila] = locale.format('%10.2f', items.coste_prod)
            hoja['P' + fila] = locale.format('%10.2f', items.venta_prod)
            hoja['AF' + fila] = items.tech

            if items.maintenance:  # El ítem tiene mantenimiento. Añadir las dos lineas correspondientes
                fila = str(curr_row + 1)
                fila_sig = str(curr_row + 2)

                hoja['A' + fila] = num_fila + 1
                hoja['A' + fila_sig] = num_fila + 2
                hoja['B' + fila_sig] = num_fila + 1
                hoja['H' + fila] = 2
                hoja['H' + fila_sig] = 20
                hoja['C' + fila] = 'Dimension Data'
                hoja['C' + fila_sig] = items.manufacturer
                hoja['E' + fila] = items.unit
                hoja['E' + fila_sig] = items.unit
                hoja['F' + fila] = items.sku_uptime
                hoja['F' + fila_sig] = items.backout_name
                hoja['G' + fila] = items.descr_uptime
                hoja['G' + fila_sig] = items.code
                hoja['I' + fila] = items.code
                hoja['I' + fila_sig] = 0
                hoja['J' + fila] = items.manufacturer
                hoja['J' + fila_sig] = ''
                hoja['K' + fila] = 1
                hoja['K' + fila_sig] = 1
                hoja['L' + fila] = 'EA'
                hoja['L' + fila_sig] = 'EA'
                hoja['M' + fila] = 'EUR'
                hoja['M' + fila_sig] = 'EUR'
                hoja['N' + fila] = 'Fixed'
                hoja['N' + fila_sig] = 'Fixed'
                hoja['O' + fila] = items.list_price_back
                hoja['O' + fila_sig] = items.list_price_back
                durac_years = int(round(items.durac/12, 0))
                hoja['P' + fila] = locale.format('%10.2f', float(items.venta_mant * durac_years))
                hoja['P' + fila_sig] = locale.format('%10.2f', float(items.venta_mant * durac_years))
                hoja['Q' + fila] = locale.format('%10.2f', float(items.cost_unit_manten * durac_years))
                hoja['Q' + fila_sig] = locale.format('%10.2f', float(items.coste_unit_back * durac_years))
                init_date_formateado = datetime.strptime(items.init_date, "%d/%m/%Y")
                end_date_formateado = datetime.strptime(items.end_date, "%d/%m/%Y")
                fecha_init = '{}/{}/{}'.format(init_date_formateado.day, init_date_formateado.month,
                                               init_date_formateado.year)
                fecha_init_limpia = '{}{}{}'.format(str(init_date_formateado.year),
                                                    str(init_date_formateado.month).zfill(2),
                                                    str(init_date_formateado.day).zfill(2))
                fecha_fin = '{}/{}/{}'.format(end_date_formateado.day, end_date_formateado.month,
                                              end_date_formateado.year)
                hoja['X' + fila] = fecha_init
                hoja['X' + fila_sig] = fecha_init
                hoja['Y' + fila] = fecha_fin
                hoja['Y' + fila_sig] = fecha_fin
                hoja['AA' + fila] = ('StartDate=' + fecha_init_limpia + '#Duration=' + str(items.durac) +
                                     '#InvoiceInterval=Yearly#InvoiceMode=anticipated')
                hoja['AA' + fila_sig] = ('StartDate=' + fecha_init_limpia + '#Duration=' + str(items.durac) +
                                         '#InvoiceInterval=Yearly#InvoiceMode=anticipated')

                hoja['AF' + fila] = items.tech
                hoja['AF' + fila_sig] = items.tech
                curr_row += 2

            curr_row += 1
            num_fila += 10

        nombre_fichero_salida_exc = 'productos_' + fabr + '.xlsx'
        nombre_fichero_salida_csv = 'productos_' + fabr + '.csv'

        fichero_salida_excel = os.path.join(directorio, nombre_fichero_salida_exc)
        fichero_salida_csv = os.path.join(directorio, nombre_fichero_salida_csv)

        todo_ok = False

        while not todo_ok:
            try:
                libro.save(fichero_salida_excel)
                todo_ok = True
            except PermissionError:
                QtWidgets.QMessageBox.warning(instance, 'Procesado de oferta',
                                              'Fichero Excel de destino {} ya abierto \n'
                                              'por favor ciérrelo y pulse Aceptar'.format(fichero_salida_excel))

        csv_from_excel(fichero_salida_excel, fichero_salida_csv, instance)
        os.remove(fichero_salida_excel)

    else:
        return


def convert_sla(text:str, instance):   # Este método convierte las descripciones de Uptime en códigos de servicio

    sla_code = instance.dict_sla.get(text, None)  # De este diccionario sacamos la equivalencia descrip --> código
    return sla_code


# def hacer_oferta_ms(lista_articulos, directorio, instance):
#
#     # En primer lugar filtramos los registros que incluyen mantenimiento
#
#     lista_ms = [item
#                 for item in lista_articulos
#                 if item.maintenance]
#
#     excel_aux = pass_to_excel(lista_ms, instance)
#     fichero_excel_out = os.path.join(directorio, 'ms.xlsx')
#     fichero_csv_out = os.path.join(directorio, 'ms.csv')
#
#     if excel_aux:
#
#         excel_aux.save(fichero_excel_out)
#         csv_from_excel(fichero_excel_out, fichero_csv_out, instance)
#         os.remove(fichero_excel_out)
